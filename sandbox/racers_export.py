import openpyxl
import requests

class Racer:
    def __init__(self, number, name, category, class_, coach, organization, weight, sex):
        self.number = number
        self.name = name
        self.category = category
        self.class_ = class_
        self.coach = coach
        self.organization = organization
        self.weight = weight
        self.sex = sex

def read_excel():
    wb = openpyxl.load_workbook('sandbox/kt_racers.xlsx')
    ws = wb['sheet1']
    racers = []
    for i in range(7,142):
        racers.append(Racer(
            ws.cell(row=i, column=1).value,
            ws.cell(row=i, column=2).value,
            ws.cell(row=i, column=3).value,
            ws.cell(row=i, column=4).value,
            ws.cell(row=i, column=5).value,
            ws.cell(row=i, column=6).value,
            ws.cell(row=i, column=7).value,
            ws.cell(row=i, column=8).value
        ))

    #напечатаем все имена из массива racers
    for racer in racers:
        print(racer.name)

#подключаемся и авторизуемся на сервере fastapi
    
if __name__ == '__main__':

    data = {    
        "username": "admin",
        "password": "admin"
    }

    response = requests.post('http://82.97.247.48:8000/token', json=data)

    if response.status_code == 200:
        result = response.json()
        # Обрабатываем полученный результат
    else:
        print(f'Ошибка при выполнении запроса: {response.status_code}')

