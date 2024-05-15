from collections import OrderedDict
import os
import openpyxl
import requests

class RacerFields:
    number = "Стартовый номер"
    fio = "ФИО"
    discipline = "Дисциплина"
    year = "Год рождения"
    birthday = "Дата рождения"
    level = "Уровень"
    coach = "Тренер"
    organization = "Клуб"
    weight1 = "Вес, кг"
    zaezd_number = "Номер заезда"
    sex = "Пол"
    
class Racer:
    def __init__(self, **kwargs):
        for field, attr in RacerFields.__dict__.items():
            if not field.startswith('_'):  #пропускаем приватные поля
                setattr(self, field, kwargs.get(attr, None))
                
        if self.birthday is not None:
            self.year = self.birthday.year
                
        self.track_number = 0

    @property
    def zaezd(self):
        if isinstance(self.zaezd_number, str):
            return self.zaezd_number
        
        return f'Заезд {self.zaezd_number}'
    
    @property
    def first_name(self):
        #Первый элмент до пробела из self.fio очищенный от ведущих и закрывающих пробелов
        return self.fio.split()[1]
    
    @property
    def last_name(self):
        return self.fio.split()[0].strip()
    
    @property
    def weight(self):
        res = self.weight1
        #проверит является ли res строкой
        if isinstance(res, str):
            res = res.split(',')[0].split('.')[0]
        return int(res)
    
    @property
    def age(self):
        return 2024 - int(self.year)
    
    @property
    def disciplines(self):
        return self.discipline
    
    def check(self):
        check_fileds = ['number', 'fio', 'year', 'discipline', 'weight1']   
        #проверить, что все поля класса заполнены
        for key in check_fileds:
            if getattr(self, key) is None:
                raise Exception(f'Не заполнено поле {key}')
        return True
    
    @classmethod
    def from_excel_row(cls, row, maps): 
        kwargs = {}
        for item in maps.items():
            #присвой значение перемнной по имени из словаря
            #item[1] может быть списком. В этом случае надо считыывать список
            if isinstance(item[1], list):
                kwargs[item[0]] = [row[i] for i in item[1] if row[i] is not None]
            else:
                kwargs[item[0]] = row[item[1]]
 
        return Racer(**kwargs)
        
class Discipline:
    name = "Дисциплина"
    distance = "Дистанция"

    def __init__(self, name, distance = 200):
        self.name = name
        self.distance = distance
        self.races = {}
        self.racers = []

    def add_racer(self, racer):
        self.racers.append(racer)
        if racer.zaezd not in self.races:
            self.races[racer.zaezd] = []
        
        self.races[racer.zaezd].append(racer)

    def race_names(self):
        return self.races.keys()
    
    def get_race(self, name):
        if name in self.races:
            return self.races[name]
        return []
    
    def values(self):
        return self.races.values()
    
    def arrange_races(self, prefix, start_num, tracks_count = 9):
        rc_qty = len(self.racers)
        zaezd_count = int(rc_qty / tracks_count)
        if zaezd_count * tracks_count < rc_qty:
            zaezd_count += 1
        
        on_track = int(rc_qty / zaezd_count)
        reminder = rc_qty - on_track * zaezd_count 

        self.races = {}
        cur_racer_idx = 0
        for z in range(0, zaezd_count):
            zaezd_name = f'{prefix} {start_num+z}' 
            self.races[zaezd_name] = []
            zaezd_racer_list = self.races[zaezd_name]
            track_num = 1
            for r in range(0, on_track):
                #вставить кортеж из номера дорожки и racer
                zaezd_racer_list.append((track_num, self.racers[cur_racer_idx]))
                cur_racer_idx += 1
                track_num += 1
            if reminder:
                zaezd_racer_list.append((track_num, self.racers[cur_racer_idx]))
                cur_racer_idx += 1
                track_num += 1
                reminder -= 1
        
        return zaezd_count
    
    def print(self):
        pass
    
class ExcelRacersFile:
    def __init__(self, filename, sheet_name, row_range = (1,100), col_range = (1,20), rearrange_races = False, tracks_count=9):
        self.filename = filename
        self.sheet_name = sheet_name
        self.row_range = row_range
        self.col_range = col_range
        self.rearrange_races = rearrange_races
        self.lines_count = tracks_count

def read_excel_header(ws):
    map = {}
    for i in range(ws.min_column, ws.max_column+1):
        value = ws.cell(row=1, column=i).value
        idx = i - 1
        if value is not None:
            if value not in map:
                map[value] = idx
            else:
                #проверить что значение в map не является списком
                if not isinstance(map[value], list):
                    map[value] = [map[value], idx]
                else:
                    map[value].append(idx)
    return map

def read_disciplines(wb):
    disciplines = OrderedDict()
    ws = wb['Дисциплины']
    #Читаем заголовок дисциплин
    map = {}
    for i in range(ws.min_column, ws.max_column+1):
        if ws.cell(row=1, column=i).value is not None:
            map[ws.cell(row=1, column=i).value] = i - 1
    #Читаем дисциплины
    for i in range(2, ws.max_row+1):
        row = [ws.cell(row=i, column=j).value for j in range(ws.min_column, ws.max_column+1)]
        discipline = Discipline(row[map[Discipline.name]], row[map[Discipline.distance]])
        disciplines[discipline.name] = discipline
        
    return disciplines

        
def read_excel(settings):
    wb = openpyxl.load_workbook(settings.filename)
    ws = wb[settings.sheet_name]
    
    disciplines = read_disciplines(wb)
    column_map = read_excel_header(ws)
    
    count = 0
    zaezd_count = 0
    for i in range(settings.row_range[0],settings.row_range[1]):
        row = [ws.cell(row=i, column=j).value for j in range(settings.col_range[0],settings.col_range[1])]
        
        racer = Racer.from_excel_row(row, column_map)
        
        if racer.number:
            count += 1
            
            if settings.rearrange_races:
                racer.zaezd_number = None
            for discipline in set(racer.disciplines):  # use set to remove duplicates and speed up the loop
                if discipline not in disciplines:
                    disciplines[discipline] = Discipline(discipline)
                disciplines[discipline].add_racer(racer)

            try:
                racer.check()
            except Exception as e:
                raise Exception(f'Ошибка в записи участника строка: {i} -  {e}')
            
            print(racer.first_name)
    
    zaezd_count = 0
    if settings.rearrange_races:
        for discipline in disciplines.values():
            zaezd_count += discipline.arrange_races('Заезд', zaezd_count+1, settings.lines_count) 

    #напечатаем все имена из массива racers
    print(f'Считано {count} участников. Дисциплин {len(disciplines)}. Заездов {zaezd_count}')
    print('Дисциплины:')
    total = 0
    for item in disciplines.keys():
        cnt = len(disciplines[item].racers)
        print(f'{item} участников {cnt}')
        total += cnt
    print(f'== Всего участников {total}')
    
    for item in disciplines:
        print(item)
        for zaezd in disciplines[item].race_names():
            z_racers = disciplines[item].get_race(zaezd)
            print(f'-{zaezd}: {len(z_racers)}')
            for i in range(0, len(z_racers)):
                racer = z_racers[i][1]
                track_number = z_racers[i][0]
                print(f'\t {track_number} {racer.number}  {racer.last_name} {racer.first_name} \t\t{racer.weight} \t{racer.age}')
    
    return disciplines            
#подключаемся и авторизуемся на сервере fastapi

# Выгружаем в Excel результат работы функции read_excel
def write_excel(filename, disciplines):
    from openpyxl.styles.alignment import Alignment
    from openpyxl.styles.fonts import Font

    #удалть filename если он уже существует
    if os.path.isfile(filename):
        os.remove(filename)

    wb = openpyxl.Workbook()
    ws = wb.active

    

    #объединяем три колонки и печтатаем заголовок "Заезды" крупным шрифтом
    ws.merge_cells('A1:C1')
    ws['A1'] = 'Заезды'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    #выравниваем по центру

    
    #ws.append(['Номер', 'ФИО', 'Год рождения', 'Дисциплина', 'вес, кг', 'Тренер', 'Организвция', 'Заезд'])
    #обвечсти ячейки чёрным цветом
    for row in ws.iter_rows():
        for cell in row:
            #обести рамкой, а не залить
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(border_style='thin', color='00000000'),
                right=openpyxl.styles.Side(border_style='thin', color='00000000'),
                top=openpyxl.styles.Side(border_style='thin', color='00000000'),
                bottom=openpyxl.styles.Side(border_style='thin', color='00000000')
                #fill=openpyxl.styles.PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
                #font=openpyxl.styles.Font(color='00000000')
                #alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center')
                #number_format='@'
            )
    
    minutes = 15
    hours = 18
    
    #Задать ширину колонок
   
    for item in disciplines:
        for zaezd in disciplines[item].race_names():
            z_racers = disciplines[item].get_race(zaezd)
            start_time = f'{hours:02d}:{minutes % 60:02d}' 
            minutes += 5
            hours = hours + minutes // 60

            ws.append(['', ''])
            ws.append(['', zaezd, start_time]) 
            #Сделать жирным
            ws[f'B{ws.max_row}'].font = Font(bold=True)
            ws.append(['', item])
            ws[f'B{ws.max_row}'].font = Font(bold=True)            
            ws.append(['Дорожка', 'ФИО', 'Год рождения', 'вес, кг'])
            last_row = ws.max_row
            for i in range(0, len(z_racers)):
                racer = z_racers[i][1]
                track_number = z_racers[i][0]
                ws.append([track_number, racer.fio, racer.year, racer.weight])
            #Обвести всё что ниже рамкой
            for row in ws.iter_rows(min_row=last_row, max_row=ws.max_row):
                for cell in row:
                    cell.border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(border_style='thin', color='00000000'),
                        right=openpyxl.styles.Side(border_style='thin', color='00000000'),
                        top=openpyxl.styles.Side(border_style='thin', color='00000000'),
                        bottom=openpyxl.styles.Side(border_style='thin', color='00000000')
                        #fill=openpyxl.styles.PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
                        #font=openpyxl.styles.Font(color='00000000')
                        #alignment=openpyxl.styles.Alignment(horizontal='center', vertical='center')
                        #number_format='@'
                    )
    wb.save(filename)

racers_file = ExcelRacersFile(
    filename ='sandbox/003_kartontol.xlsx', 
    sheet_name='Участники',  
    row_range=(2,100), 
    col_range=(1,20),
    rearrange_races=True,
    tracks_count=6)
    
if __name__ == '__main__':
    try:
        disciplines = read_excel(racers_file)
        write_excel('sandbox/003_kartontol_out.xlsx', disciplines)
    except Exception as e:
        print(f'Ошибка: {e}')